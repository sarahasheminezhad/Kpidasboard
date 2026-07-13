from bs4 import BeautifulSoup
from openpyxl import Workbook

input_file = "shapino.html"

with open(input_file, "r", encoding="utf-8") as f:
    soup = BeautifulSoup(f, "html.parser")

# پیدا کردن همه تگ‌هایی که کلاس WylIhj دارند
elements = soup.find_all(class_="svelte-158jj44")

# استخراج متن
data = [el.get_text(strip=True) for el in elements]

# ذخیره در اکسل
wb = Workbook()
ws = wb.active
ws.title = "svelte-158jj44"

ws["A1"] = "Text"

for i, text in enumerate(data, start=2):
    ws[f"A{i}"] = text

wb.save("shapino.xlsx")

print("Done.")